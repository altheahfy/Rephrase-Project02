#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
例文入力元.xlsxの自動チェックスクリプト
- イラスト未作成の表現を検出
- slottext.jsonに未登録の表現を検出
- 例文入力元.xlsxの右側空白列に自動でリストアップ

作成者: GitHub Copilot
作成日: 2025年7月20日
"""

import pandas as pd
import json
import os
import re
from pathlib import Path
from typing import Set, List, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill

class MissingAssetsChecker:
    def __init__(self):
        # ファイルパス設定
        self.excel_file = "例文入力元.xlsx"
        self.slottext_file = "slottext.json"
        self.image_meta_file = "../完全トレーニングUI完成フェーズ３/project-root/Rephrase-Project/training/image_meta_tags.json"
        self.image_folder = "../完全トレーニングUI完成フェーズ３/project-root/Rephrase-Project/training/slot_images/common"
        
        # データ読み込み
        self.df = None
        self.slottext_rules = []
        self.image_meta_tags = []
        self.image_files = set()
        
        # 結果保存用
        self.missing_assets = []  # 統合されたアセットリスト
        
    def load_data(self):
        """各データファイルを読み込み"""
        print("📊 データ読み込み中...")
        
        # Excel読み込み
        if os.path.exists(self.excel_file):
            self.df = pd.read_excel(self.excel_file)
            print(f"✅ Excel読み込み完了: {len(self.df)}行")
        else:
            print(f"❌ Excel файल не найден: {self.excel_file}")
            return False
            
        # slottext.json読み込み
        if os.path.exists(self.slottext_file):
            with open(self.slottext_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.slottext_rules = data.get('rules', [])
            print(f"✅ slottext.json読み込み完了: {len(self.slottext_rules)}ルール")
        else:
            print(f"❌ slottext.json не найден: {self.slottext_file}")
            return False
            
        # image_meta_tags.json読み込み
        if os.path.exists(self.image_meta_file):
            with open(self.image_meta_file, 'r', encoding='utf-8') as f:
                self.image_meta_tags = json.load(f)
            print(f"✅ image_meta_tags.json読み込み完了: {len(self.image_meta_tags)}項目")
        else:
            print(f"❌ image_meta_tags.json не найден: {self.image_meta_file}")
            return False
            
        # 画像ファイル一覧取得
        if os.path.exists(self.image_folder):
            image_path = Path(self.image_folder)
            self.image_files = {f.stem.lower() for f in image_path.glob("*.png")}
            print(f"✅ 画像ファイル読み込み完了: {len(self.image_files)}ファイル")
        else:
            print(f"❌ 画像フォルダが見つかりません: {self.image_folder}")
            return False
            
        return True
    
    def extract_words_from_text(self, text: str) -> Set[str]:
        """テキストから単語を抽出（英単語のみ）"""
        if pd.isna(text) or not isinstance(text, str):
            return set()
            
        # 検出対象外とする一般的な単語
        excluded_words = {
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by',
            'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did',
            'will', 'would', 'can', 'could', 'may', 'might', 'must', 'should', 'shall',
            'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her', 'us', 'them',
            'my', 'your', 'his', 'her', 'its', 'our', 'their', 'this', 'that', 'these', 'those'
        }
        
        # 英単語のみを抽出（アルファベット2文字以上）
        words = re.findall(r'\b[a-zA-Z]{2,}\b', text.lower())
        # 除外単語を取り除く
        return set(word for word in words if word not in excluded_words)
    
    def check_slottext_coverage(self, text: str) -> bool:
        """slottext.jsonでカバーされているかチェック"""
        if pd.isna(text) or not isinstance(text, str):
            return True
            
        # 一般的すぎるパターンを除外（具体的な単語のslottextニーズを見落とさないため）
        excluded_general_patterns = [
            "\\b\\w+ed\\b",   # 一般的な過去形パターン
            "\\b\\w+ing\\b",  # 一般的な現在分詞パターン
            "\\b\\w+s\\b",    # 一般的な複数形パターン
        ]
        
        # 各ルールに対してマッチングチェック
        for rule in self.slottext_rules:
            condition = rule.get('condition', '')
            
            # 一般的すぎるパターンをスキップ
            if condition in excluded_general_patterns:
                continue
                
            try:
                if re.search(condition, text, re.IGNORECASE):
                    return True
            except re.error:
                # 正規表現エラーの場合はスキップ
                continue
                
        return False
    
    def check_image_coverage(self, text: str) -> bool:
        """画像でカバーされているかチェック"""
        if pd.isna(text) or not isinstance(text, str):
            return True
            
        words = self.extract_words_from_text(text)
        
        # 画像メタタグでのチェック
        for meta in self.image_meta_tags:
            meta_tags = meta.get('meta_tags', [])
            for tag in meta_tags:
                if tag.lower() in words:
                    return True
                    
        # 直接ファイル名でのチェック
        for word in words:
            if word in self.image_files:
                return True
                
        return False
    
    def analyze_excel_data(self):
        """Excelデータを分析して未作成アセットを特定"""
        print("🔍 未作成アセット分析中...")
        
        # 分析対象列
        text_columns = ['SlotPhrase', 'SubslotElement']
        
        missing_assets_set = set()  # 統合された未作成アセットセット
        
        for _, row in self.df.iterrows():
            for col in text_columns:
                text = row.get(col)
                if pd.isna(text) or not isinstance(text, str):
                    continue
                    
                text = text.strip()
                if not text:
                    continue
                
                # 個別単語レベルでチェック
                words = self.extract_words_from_text(text)
                for word in words:
                    # 単語にイラストもslottextもない場合のみ未作成アセットとして追加
                    has_image = self.check_word_image_coverage(word)
                    has_slottext = self.check_slottext_coverage(word)
                    
                    if not has_image and not has_slottext:
                        # どちらも存在しない場合、イラスト候補として追加
                        missing_assets_set.add(word)
        
        self.missing_assets = sorted(list(missing_assets_set))
        
        print(f"🔍 未作成アセット合計: {len(self.missing_assets)}項目")
    
    def check_word_image_coverage(self, word: str) -> bool:
        """単語がイラストでカバーされているかチェック"""
        word_lower = word.lower()
        
        # チェック対象の単語リスト（元の単語 + 単数形候補）
        words_to_check = [word_lower]
        
        # 複数形を単数形に変換してチェック対象に追加
        if word_lower.endswith('s') and len(word_lower) > 2:
            # 基本的な複数形のルール
            if word_lower.endswith('ies'):
                # cities -> city, stories -> story
                singular = word_lower[:-3] + 'y'
                words_to_check.append(singular)
            elif word_lower.endswith('es') and word_lower[-3] in 'sxz':
                # boxes -> box, glasses -> glass
                singular = word_lower[:-2]
                words_to_check.append(singular)
            elif word_lower.endswith('es') and word_lower.endswith(('ches', 'shes')):
                # watches -> watch, dishes -> dish
                singular = word_lower[:-2]
                words_to_check.append(singular)
            else:
                # 通常の複数形: days -> day, cats -> cat
                singular = word_lower[:-1]
                words_to_check.append(singular)
        
        # 各候補単語でチェック
        for check_word in words_to_check:
            # 画像メタタグでのチェック
            for meta in self.image_meta_tags:
                meta_tags = meta.get('meta_tags', [])
                if check_word in [tag.lower() for tag in meta_tags]:
                    return True
                    
            # 直接ファイル名でのチェック
            if check_word in self.image_files:
                return True
                
        return False
    
    def update_excel_with_results(self):
        """結果をExcelファイルに反映"""
        print("📝 Excelファイル更新中...")
        
        # openpyxlでExcelファイルを開く
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active
        
        # スタイル定義
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        missing_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # ヘッダー行の設定（行1）
        ws.cell(row=1, column=22, value="未作成アセット").font = header_font
        ws.cell(row=1, column=22).fill = header_fill
        
        # データ行の処理
        for idx, row in self.df.iterrows():
            excel_row = idx + 2  # Excelは1-indexedで、ヘッダー行があるので+2
            
            # 未作成アセット収集
            missing_assets_for_row = []
            
            for col in ['SlotPhrase', 'SubslotElement']:
                text = row.get(col)
                if pd.isna(text) or not isinstance(text, str):
                    continue
                    
                text = text.strip()
                if not text:
                    continue
                
                # 個別単語レベルでチェック
                words = self.extract_words_from_text(text)
                for word in words:
                    # 単語にイラストもslottextもない場合のみ未作成アセットとして追加
                    has_image = self.check_word_image_coverage(word)
                    has_slottext = self.check_slottext_coverage(word)
                    
                    if not has_image and not has_slottext:
                        # どちらも存在しない場合、イラスト候補として追加
                        missing_assets_for_row.append(word)
            
            # 結果をセルに書き込み
            if missing_assets_for_row:
                # 重複を除去してソート
                unique_assets = sorted(list(set(missing_assets_for_row)))
                cell = ws.cell(row=excel_row, column=22, value=", ".join(unique_assets))
                cell.fill = missing_fill
        
        # ファイル保存
        output_file = "例文入力元_チェック結果.xlsx"
        wb.save(output_file)
        print(f"✅ 結果を保存しました: {output_file}")
        
        return output_file
    
    def print_summary(self):
        """結果のサマリーを表示"""
        print("\n" + "="*60)
        print("📊 未作成アセット分析結果")
        print("="*60)
        
        print(f"\n🔍 未作成アセット ({len(self.missing_assets)}個):")
        print("  📷 = イラスト候補, 📝 = slottext候補")
        if self.missing_assets:
            for i, item in enumerate(self.missing_assets[:30], 1):  # 最初の30個を表示
                print(f"  {i:2d}. {item}")
            if len(self.missing_assets) > 30:
                print(f"     ... 他{len(self.missing_assets) - 30}項目")
        else:
            print("  なし")
            
        print("\n" + "="*60)
    
    def run(self):
        """メイン実行関数"""
        print("🚀 例文入力元.xlsx 自動チェック開始")
        print("="*60)
        
        # データ読み込み
        if not self.load_data():
            print("❌ データ読み込みに失敗しました")
            return False
        
        # 分析実行
        self.analyze_excel_data()
        
        # 結果表示
        self.print_summary()
        
        # Excel更新
        output_file = self.update_excel_with_results()
        
        print(f"\n✅ 処理完了！結果ファイル: {output_file}")
        print("💡 次のステップ:")
        print("   1. 結果ファイルを開いて、赤くハイライトされた項目を確認")
        print("   2. 📷マークは画像作成候補、📝マークはslottext追加候補")
        print("   3. 実際にイラストかslottextかは内容を見て手作業で判断")
        
        return True

def main():
    """メイン実行関数"""
    checker = MissingAssetsChecker()
    success = checker.run()
    
    if success:
        print("\n🎉 自動チェック完了")
    else:
        print("\n💥 エラーが発生しました")

if __name__ == "__main__":
    main()
