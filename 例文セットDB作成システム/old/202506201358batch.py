
import json
import openpyxl
import re
import sys

# slottext.json 読み込み
with open("slottext.json", "r", encoding="utf-8") as f:
    slottext_rules = json.load(f)["rules"]

def infer_slot_text(phrase, slot_name=None, subslots=None, idx=None):
    phrase_lc = phrase.lower()
    matches = []
    had_to_matched = False

    for rule in slottext_rules:
        condition = rule["condition"]
        guidance = rule["guidance"]
        m = re.search(condition, phrase_lc, re.IGNORECASE)
        if m:
            if condition == r"\b\w+ed\b":
                skip_past = False
                if subslots is not None and idx is not None:
                    if idx > 1:
                        prev_phrase1 = subslots[idx - 2].get("phrase", "").lower()
                        if prev_phrase1 in ["have", "has", "had"]:
                            skip_past = True
                    if idx > 2:
                        prev_phrase2 = subslots[idx - 3].get("phrase", "").lower()
                        if prev_phrase2 in ["have", "has", "had"]:
                            skip_past = True
                if skip_past:
                    continue
            if condition == "^had to":
                had_to_matched = True
                matches.append((m.start(), guidance))
                continue
            if condition == "^had" and had_to_matched:
                continue
            matches.append((m.start(), guidance))

    matches.sort(key=lambda x: x[0])
    texts = [g for _, g in matches]

    if "be動詞過去" in texts and re.search(r"^was\s+\w+ing", phrase_lc):
        texts = [t for t in texts if t != "be動詞過去"]

    if slot_name not in ["V", "sub-v"]:
        texts = [t for t in texts if t != "過去形"]

    if slot_name in ["V", "sub-v"] and subslots is not None and idx is not None:
        helpers = []
        if idx > 1:
            helpers.append(subslots[idx - 2].get("phrase", "").lower())
        if idx > 2:
            helpers.append(subslots[idx - 3].get("phrase", "").lower())

        if any(h in ["have", "has"] for h in helpers):
            texts = [t for t in texts if t != "過去形"]
            texts.append("現在完了")
        elif any(h == "had" for h in helpers):
            texts = [t for t in texts if t != "過去形"]
            texts.append("過去完了")

    if slot_name in ["V", "sub-v"] and len(phrase.split()) >= 2:
        texts.append("句動詞")

    return "、".join(texts)

def safe_strip(value):
    return value.strip() if isinstance(value, str) else ""

input_file = sys.argv[1] if len(sys.argv) > 1 else "例文入力元.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb.active

header = [cell.value for cell in ws[1]]
data_rows = list(ws.iter_rows(min_row=2, values_only=True))

output_data = []
subslots_accum = [{"phrase": (row[header.index("SubslotElement")] or "")} for row in data_rows]

for idx, row in enumerate(data_rows, start=1):
    row_dict = dict(zip(header, row))
    構文ID = row_dict.get("構文ID") or ""
    v_group_key = safe_strip(row_dict.get("V_group_key"))
    example_id = safe_strip(row_dict.get("例文ID"))
    slot = safe_strip(row_dict.get("Slot"))
    slot_phrase = safe_strip(row_dict.get("SlotPhrase"))
    phrase_type = safe_strip(row_dict.get("PhraseType"))
    subslot_id = safe_strip(row_dict.get("SubslotID"))
    subslot_element = safe_strip(row_dict.get("SubslotElement"))
    slot_display_order = row_dict.get("Slot_display_order") or 0
    display_order = row_dict.get("display_order") or 0

    slot_text = infer_slot_text(slot_phrase, slot_name=slot, subslots=subslots_accum, idx=idx) if slot_phrase else ""
    subslot_text = infer_slot_text(subslot_element, slot_name=subslot_id, subslots=subslots_accum, idx=idx) if subslot_element else ""

    output_data.append({
        "構文ID": 構文ID,
        "V_group_key": v_group_key,
        "例文ID": example_id,
        "Slot": slot,
        "SlotPhrase": slot_phrase,
        "SlotText": slot_text,
        "PhraseType": phrase_type,
        "SubslotID": subslot_id,
        "SubslotElement": subslot_element,
        "SubslotText": subslot_text,
        "Slot_display_order": slot_display_order,
        "display_order": display_order
    })

with open("slot_order_data.json", "w", encoding="utf-8") as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"✅ slot_order_data.json を出力しました。レコード数: {len(output_data)}")
