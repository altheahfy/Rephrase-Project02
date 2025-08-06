
import json
import openpyxl
import re
import sys

# ✅ JSONファイル指定（コマンド引数対応）
json_file = sys.argv[1] if len(sys.argv) > 1 else "example_3.json"
with open(json_file, "r", encoding="utf-8") as f:
    data = json.load(f)

with open("slottext.json", "r", encoding="utf-8") as f:
    slottext_rules = json.load(f)["rules"]

# ✅ 補助テキスト自動補完
def infer_slot_text(phrase):
    """
    補助テキストを推定する関数。
    すべて正規表現評価に統一。
    """
    phrase_lc = phrase.lower()
    for rule in slottext_rules:
        condition = rule["condition"]
        guidance = rule["guidance"]
        if re.search(condition, phrase_lc, re.IGNORECASE):
            return guidance
    return ""

# ✅ Excelロード
wb = openpyxl.load_workbook("DB_know_ex001.xlsx")

# ✅ 各シート処理（元のまま維持）
master_sheet = wb["example_master"]
構文ID = data.get("構文ID", "")
V_group_key = data.get("V_group_key", "")
example_id = data.get("例文ID", "")  # 修正: 日本語フィールドに対応
original = data.get("原文", "")  # 修正: 日本語フィールドに対応

values = [構文ID or "", V_group_key or "", example_id or "", original or ""]
master_sheet.append(values)

phrase_sheet = wb["slot_phrase_pool"]
existing_pairs = set(((row[1].value or "").strip(), (row[2].value or "").strip()) for row in phrase_sheet.iter_rows(min_row=2))
for slot in data["slot_data"]:
    key = (slot["slot"].strip(), slot["phrase"].strip())
    if key not in existing_pairs:
        phrase_sheet.append([V_group_key, slot["slot"], slot["phrase"]])
        existing_pairs.add(key)

type_sheet = wb["phrase_type_rules"]
for slot in data["slot_data"]:
    type_sheet.append([slot["phrase"], slot["type"]])

text_sheet = wb["slottext_rules"]
for slot in data["slot_data"]:
    if slot["type"] == "word":
        slot_text = slot.get("slot_text", "") or infer_slot_text(slot["phrase"])
        text_sheet.append([slot["phrase"], slot_text])

structure_sheet = wb["slot_structure"]
for slot in data["slot_data"]:
    structure_sheet.append([slot["phrase"], slot["slot"], slot["type"], slot.get("slot_text", "")])

mapping_sheet = wb["subslot_mapping"]
substructure_sheet = wb["subslot_structure"]
for slot in data["slot_data"]:
    if slot.get("type") == "clause" and "subslots" in slot:
        for sub in slot["subslots"]:
            mapping_sheet.append([
                slot.get("phrase", ""),
                sub.get("slot", ""),
                sub.get("phrase", "")
            ])

# ✅ 差分追加: 正しい subslot_structure への記載処理
existing_subslots = set((row[0].value.strip(), row[1].value.strip(), row[2].value.strip()) 
                        for row in substructure_sheet.iter_rows(min_row=2))
current_parent_phrases = set(slot.get("phrase", "").strip() for slot in data["slot_data"])

for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
    parent_phrase, sub_slot, sub_phrase = row
    if parent_phrase.strip() not in current_parent_phrases:
        continue
    key = (V_group_key, parent_phrase.strip(), sub_slot.strip(), sub_phrase.strip())
    if key[:3] not in existing_subslots:
        sub_text = infer_slot_text(sub_phrase.strip())
        substructure_sheet.append([
            V_group_key,
            parent_phrase.strip(),
            sub_slot.strip(),
            sub_phrase.strip(),
            sub_text
        ])

# 保存
output_path = "DB_know_ex001_updated.xlsx"
wb.save(output_path)

with open("log_output.txt", "w", encoding="utf-8") as log:
    log.write(f"✅ DB updated and saved to: {output_path}\n")
    log.write(f"Written master row: {values}\n")
