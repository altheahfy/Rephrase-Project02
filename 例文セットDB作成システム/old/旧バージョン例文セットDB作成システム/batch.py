
import json
import openpyxl
import re
import sys

json_file = sys.argv[1] if len(sys.argv) > 1 else "example_3.json"
with open(json_file, "r", encoding="utf-8") as f:
    data = json.load(f)

with open("slottext.json", "r", encoding="utf-8") as f:
    slottext_rules = json.load(f)["rules"]

if "slot_data" in data:
    for idx, slot in enumerate(data["slot_data"], start=1):
        slot["slot_display_order"] = idx
else:
    with open("log_output.txt", "w", encoding="utf-8") as log:
        log.write("❌ slot_data が存在しません。処理を中断しました。\n")
    sys.exit(1)

def infer_slot_text(phrase, slot_name=None, subslots=None, idx=None):
    phrase_lc = phrase.lower()
    matches = []
    had_to_matched = False

    for rule in slottext_rules:
        condition = rule["condition"]
        guidance = rule["guidance"]
        m = re.search(condition, phrase_lc, re.IGNORECASE)
        if m:
            if condition == r"\b\w+ed\b":
                skip_past = False
                if subslots is not None and idx is not None:
                    if idx > 1:
                        prev_phrase1 = subslots[idx - 2].get("phrase", "").lower()
                        if prev_phrase1 in ["have", "has", "had"]:
                            skip_past = True
                    if idx > 2:
                        prev_phrase2 = subslots[idx - 3].get("phrase", "").lower()
                        if prev_phrase2 in ["have", "has", "had"]:
                            skip_past = True
                if skip_past:
                    continue
            if condition == "^had to":
                had_to_matched = True
                matches.append((m.start(), guidance))
                continue
            if condition == "^had" and had_to_matched:
                continue
            matches.append((m.start(), guidance))

    matches.sort(key=lambda x: x[0])
    texts = [g for _, g in matches]

    if "be動詞過去" in texts and re.search(r"^was\s+\w+ing", phrase_lc):
        texts = [t for t in texts if t != "be動詞過去"]

    if slot_name not in ["V", "sub-v"]:
        texts = [t for t in texts if t != "過去形"]

    if slot_name in ["V", "sub-v"] and subslots is not None and idx is not None:
        helpers = []
        if idx > 1:
            helpers.append(subslots[idx - 2].get("phrase", "").lower())
        if idx > 2:
            helpers.append(subslots[idx - 3].get("phrase", "").lower())

        if any(h in ["have", "has"] for h in helpers):
            texts = [t for t in texts if t != "過去形"]
            texts.append("現在完了")
        elif any(h == "had" for h in helpers):
            texts = [t for t in texts if t != "過去形"]
            texts.append("過去完了")

    # 複数語のVまたはsub-vの場合「句動詞」を追加
    if slot_name in ["V", "sub-v"] and len(phrase.split()) >= 2:
        texts.append("句動詞")

    return "、".join(texts)

wb = openpyxl.load_workbook("DB_know_ex001.xlsx")
master_sheet = wb["example_master"]
phrase_sheet = wb["slot_phrase_pool"]
type_sheet = wb["phrase_type_rules"]
text_sheet = wb["slottext_rules"]
structure_sheet = wb["slot_structure"]
mapping_sheet = wb["subslot_mapping"]
substructure_sheet = wb["subslot_structure"]

構文ID = data.get("構文ID", "")
V_group_key = data.get("V_group_key", "")
example_id = data.get("例文ID", "")
original = data.get("原文", "")
values = [構文ID or "", V_group_key or "", example_id or "", original or ""]
master_sheet.append(values)

existing_pairs = set(((row[1].value or "").strip(), (row[2].value or "").strip()) for row in phrase_sheet.iter_rows(min_row=2))
for slot in data["slot_data"]:
    key = (slot["slot"].strip(), slot["phrase"].strip())
    if key not in existing_pairs:
        phrase_sheet.append([V_group_key, slot["slot"], slot["phrase"]])
        existing_pairs.add(key)

for slot in data["slot_data"]:
    type_sheet.append([slot["phrase"], slot["type"]])

for slot in data["slot_data"]:
    if slot["type"] == "word":
        slot_text = slot.get("slot_text", "") or infer_slot_text(slot["phrase"], slot_name=slot["slot"])
        text_sheet.append([slot["phrase"], slot_text])

header = [cell.value for cell in structure_sheet[1]]
header_lc = [cell.lower() if cell else "" for cell in header]
if 'slot_display_order' in header_lc:
    order_col_idx = header_lc.index('slot_display_order') + 1
else:
    with open("log_output.txt", "w", encoding="utf-8") as log:
        log.write("❌ slot_display_order 列が見つかりません（大文字小文字無視比較後）。処理を中断しました。\n")
    sys.exit(1)

next_row_idx = structure_sheet.max_row + 1

for slot in data["slot_data"]:
    structure_sheet.append([
        slot["phrase"],
        slot["slot"],
        slot["type"],
        slot.get("slot_text", "") or (infer_slot_text(slot["phrase"], slot_name=slot["slot"]) if slot["type"] == "word" else "")
    ])
    structure_sheet.cell(row=next_row_idx, column=order_col_idx).value = slot.get("slot_display_order", "")
    next_row_idx += 1

for slot in data["slot_data"]:
    if slot.get("type") == "clause" and "subslots" in slot:
        for idx, sub in enumerate(slot["subslots"], start=1):
            mapping_sheet.append([
                slot.get("phrase", ""),
                sub.get("slot", ""),
                sub.get("phrase", "")
            ])

            substructure_sheet.append([
                V_group_key,
                slot.get("slot", ""),
                sub.get("slot", ""),
                sub.get("phrase", ""),
                infer_slot_text(sub.get("phrase", ""), slot_name=sub.get("slot", ""), subslots=slot["subslots"], idx=idx),
                idx
            ])

wb.save("DB_know_ex001_updated.xlsx")
