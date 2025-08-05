
import json
import openpyxl
import re
import sys

# slottext.json 読み込み
with open("slottext.json", "r", encoding="utf-8") as f:
    slottext_rules = json.load(f)["rules"]

def infer_slot_text(phrase, slot_name=None, subslots=None, idx=None):
    phrase_lc = phrase.lower()
    matches = []
    had_to_matched = False

    for rule in slottext_rules:
        condition = rule["condition"]
        guidance = rule["guidance"]
        m = re.search(condition, phrase_lc, re.IGNORECASE)
        if m:
            matches.append((m.start(), guidance))

    matches.sort(key=lambda x: x[0])
    texts = [g for _, g in matches]

    # 特例処理：V, sub-v の複数語 → 句動詞
    if slot_name in ["V", "sub-v"] and len(phrase.split()) >= 2:
        texts.append("句動詞")

    return "、".join(texts)

# Excel 読み込み
input_file = sys.argv[1] if len(sys.argv) > 1 else "例文入力フォーマット.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb.active

header = [cell.value for cell in ws[1]]
data_rows = list(ws.iter_rows(min_row=2, values_only=True))

# 出力データ構造
output_data = []

for row in data_rows:
    row_dict = dict(zip(header, row))
    slot = row_dict.get("Slot", "").strip()
    slot_phrase = row_dict.get("SlotPhrase", "").strip() or ""
    phrase_type = row_dict.get("PhraseType", "").strip() or ""
    subslot_id = row_dict.get("SubslotID", "").strip() or ""
    subslot_element = row_dict.get("SubslotElement", "").strip() or ""
    v_group_key = row_dict.get("V_group_key", "").strip()
    example_id = row_dict.get("例文ID", "").strip()
    slot_display_order = row_dict.get("Slot_display_order", 0) or 0
    display_order = row_dict.get("display_order", 0) or 0

    slot_text = ""
    subslot_text = ""
    if slot_phrase:
        slot_text = infer_slot_text(slot_phrase, slot_name=slot)
    if subslot_element:
        subslot_text = infer_slot_text(subslot_element, slot_name=subslot_id)

    output_data.append({
        "V_group_key": v_group_key,
        "例文ID": example_id,
        "Slot": slot,
        "SlotPhrase": slot_phrase,
        "SlotText": slot_text,
        "PhraseType": phrase_type,
        "SubslotID": subslot_id,
        "SubslotElement": subslot_element,
        "SubslotText": subslot_text,
        "Slot_display_order": slot_display_order,
        "display_order": display_order
    })

# JSON 出力
with open("slot_order_data.json", "w", encoding="utf-8") as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print("✅ slot_order_data.json を出力しました。")
