
import json
import openpyxl
import re
import sys

with open("slottext.json", "r", encoding="utf-8") as f:
    slottext_rules = json.load(f)["rules"]

def infer_slot_text(phrase, slot_name=None, subslots=None, idx=None):
    phrase_lc = phrase.lower()
    matches = []
    had_to_matched = False

    for rule in slottext_rules:
        condition = rule["condition"]
        guidance = rule["guidance"]
        m = re.search(condition, phrase_lc, re.IGNORECASE)
        if m:
            if guidance == "進行形のbe動詞" and slot_name not in ["Aux", "sub-aux"]:
                continue
            matches.append((m.start(), guidance))

    if slot_name in ["V", "sub-v"]:
        if re.match(r"^is$|^are$", phrase_lc):
            matches.append((0, "be動詞"))
        if re.match(r"^was$|^were$", phrase_lc):
            matches.append((0, "be動詞過去"))

    if slot_name in ["Aux", "sub-aux"] and subslots is not None and idx is not None:
        next_phrases = []
        if idx < len(subslots):
            next_phrases.append(subslots[idx].get("phrase", "").lower())
        if idx + 1 < len(subslots):
            next_phrases.append(subslots[idx + 1].get("phrase", "").lower())
        if any(re.search(r"ing$", p) for p in next_phrases):
            if re.match(r"^is$|^are$", phrase_lc):
                matches.append((0, "現在進行形"))
            if re.match(r"^was$|^were$", phrase_lc):
                matches.append((0, "過去進行形"))

    matches.sort(key=lambda x: x[0])
    texts = []
    seen = set()
    for _, g in matches:
        if g not in seen:
            texts.append(g)
            seen.add(g)

    return "、".join(texts)

def safe_strip(value):
    return value.strip() if isinstance(value, str) else ""

input_file = sys.argv[1] if len(sys.argv) > 1 else "例文入力元.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb.active

header = [cell.value for cell in ws[1]]
data_rows = list(ws.iter_rows(min_row=2, values_only=True))

output_data = []
subslots_accum = [{"phrase": (row[header.index("SubslotElement")] or "")} for row in data_rows]

for idx, row in enumerate(data_rows, start=1):
    row_dict = dict(zip(header, row))
    構文ID = row_dict.get("構文ID") or ""
    v_group_key = safe_strip(row_dict.get("V_group_key"))
    example_id = safe_strip(row_dict.get("例文ID"))
    slot = safe_strip(row_dict.get("Slot"))
    slot_phrase = safe_strip(row_dict.get("SlotPhrase"))
    phrase_type = safe_strip(row_dict.get("PhraseType"))
    subslot_id = safe_strip(row_dict.get("SubslotID"))
    subslot_element = safe_strip(row_dict.get("SubslotElement"))
    slot_display_order = row_dict.get("Slot_display_order") or 0
    display_order = row_dict.get("display_order") or 0

    slot_text = infer_slot_text(slot_phrase, slot_name=slot, subslots=subslots_accum, idx=idx) if slot_phrase else ""
    subslot_text = infer_slot_text(subslot_element, slot_name=subslot_id, subslots=subslots_accum, idx=idx) if subslot_element else ""

    output_data.append({
        "構文ID": 構文ID,
        "V_group_key": v_group_key,
        "例文ID": example_id,
        "Slot": slot,
        "SlotPhrase": slot_phrase,
        "SlotText": slot_text,
        "PhraseType": phrase_type,
        "SubslotID": subslot_id,
        "SubslotElement": subslot_element,
        "SubslotText": subslot_text,
        "Slot_display_order": slot_display_order,
        "display_order": display_order
    })

with open("slot_order_data.json", "w", encoding="utf-8") as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"✅ slot_order_data.json を出力しました。レコード数: {len(output_data)}")
